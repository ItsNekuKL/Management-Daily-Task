"""
ekspor_excel.py — Mengekspor daftar tugas ke file Excel (.xlsx) dengan format yang rapi dan profesional.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from function.ui import tampilkan_pesan


def _buat_border_tipis():
    sisi = Side(style="thin")
    return Border(left=sisi, right=sisi, top=sisi, bottom=sisi)


def _buat_border_medium():
    sisi = Side(style="medium")
    return Border(left=sisi, right=sisi, top=sisi, bottom=sisi)


def ekspor_excel(daftar_tugas):
    print(f"\n═══ EKSPOR KE EXCEL ═══")

    if not daftar_tugas:
        tampilkan_pesan("Tidak ada tugas untuk diekspor.", "warning")
        return

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nama_file = f"daftar_tugas_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Tugas"

    # Judul
    ws.merge_cells("A1:F1")
    sel_judul = ws["A1"]
    sel_judul.value = "DAFTAR TUGAS HARIAN"
    sel_judul.font = Font(name="Calibri", size=16, bold=True)
    sel_judul.alignment = Alignment(horizontal="center", vertical="center")
    sel_judul.border = _buat_border_medium()
    ws.row_dimensions[1].height = 36

    # Sub judul
    ws.merge_cells("A2:F2")
    sel_sub = ws["A2"]
    sel_sub.value = f"Diekspor pada: {datetime.datetime.now().strftime('%d %B %Y, %H:%M:%S')}"
    sel_sub.font = Font(name="Calibri", size=10, italic=True)
    sel_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # Header
    header_cols = [
        ("A4", "No.", 6),
        ("B4", "Nama Tugas", 36),
        ("C4", "Status", 14),
        ("D4", "Ditambahkan", 22),
        ("E4", "Diselesaikan", 22),
        ("F4", "Ket.", 12),
    ]

    for cell_ref, label, lebar in header_cols:
        col_letter = cell_ref[0]
        sel = ws[cell_ref]
        sel.value = label
        sel.font = Font(name="Calibri", size=11, bold=True)
        sel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sel.border = _buat_border_tipis()
        ws.column_dimensions[col_letter].width = lebar

    ws.row_dimensions[4].height = 28

    # Data
    border_tipis = _buat_border_tipis()

    for i, tugas in enumerate(daftar_tugas):
        baris = i + 5
        selesai = tugas.get("selesai", False)

        status_teks = "SELESAI" if selesai else "Belum"
        waktu_selesai = tugas.get("waktu_selesai", "-")
        ket = "Tuntas" if selesai else "On Going"

        data_baris = [
            (f"A{baris}", i + 1, "center", False),
            (f"B{baris}", tugas["nama"], "left", False),
            (f"C{baris}", status_teks, "center", True),
            (f"D{baris}", tugas.get("waktu_dibuat", "-"), "center", False),
            (f"E{baris}", waktu_selesai, "center", False),
            (f"F{baris}", ket, "center", True),
        ]

        for ref, val, align, bold in data_baris:
            sel = ws[ref]
            sel.value = val
            sel.font = Font(name="Calibri", size=10, bold=bold)
            sel.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            sel.border = border_tipis

        ws.row_dimensions[baris].height = 20

    # Ringkasan
    total_baris = len(daftar_tugas) + 5
    ws.row_dimensions[total_baris].height = 22

    jumlah_selesai = sum(1 for t in daftar_tugas if t.get("selesai"))
    jumlah_belum = len(daftar_tugas) - jumlah_selesai

    ws.merge_cells(f"A{total_baris}:B{total_baris}")
    ws[f"A{total_baris}"].value = "RINGKASAN"
    ws[f"A{total_baris}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{total_baris}"].alignment = Alignment(horizontal="center", vertical="center")

    ringkasan = [
        (f"C{total_baris}", f"Total: {len(daftar_tugas)}"),
        (f"D{total_baris}", f"Selesai: {jumlah_selesai}"),
        (f"E{total_baris}", f"Belum: {jumlah_belum}"),
        (f"F{total_baris}", f"={jumlah_selesai}/{len(daftar_tugas)}"),
    ]

    for ref, val in ringkasan:
        sel = ws[ref]
        sel.value = val
        sel.font = Font(name="Calibri", size=10, bold=True)
        sel.alignment = Alignment(horizontal="center", vertical="center")
        sel.border = border_tipis

    ws.freeze_panes = "A5"

    wb.save(nama_file)
    tampilkan_pesan(f"File Excel berhasil dibuat: {nama_file}", "sukses")
    return nama_file